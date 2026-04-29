"""Factory helpers for building in-memory XLSX workbooks in tests.

Usage::

    from tests.fixtures.import_fixtures import (
        build_workbook,
        make_loan_row,
        make_balance_row,
        minimal_valid_workbook,
    )

    wb_bytes = minimal_valid_workbook()          # happy-path 3-sheet file
    custom   = build_workbook({                  # fully custom
        "Settings": [make_setting_row()],
        "Loans":    [make_loan_row(code="my_loan")],
        "Balances": [make_balance_row(loan_code="my_loan")],
    })
"""

from __future__ import annotations

from io import BytesIO

from openpyxl import Workbook

from app.services.import_.header_validator import SHEET_COLUMNS

# ---------------------------------------------------------------------------
# Low-level workbook builder
# ---------------------------------------------------------------------------


def build_workbook(sheets: dict[str, list[dict]]) -> bytes:
    """Create an XLSX file in memory from *sheets*.

    Each key is the sheet name, each value is a list of row-dicts whose
    keys must match the columns defined in ``SHEET_COLUMNS``.

    Returns raw ``bytes`` ready for ``openpyxl.load_workbook(BytesIO(...))``.
    """
    wb = Workbook()
    # Remove default "Sheet" created by openpyxl
    wb.remove(wb.active)  # type: ignore[arg-type]

    for sheet_name, rows in sheets.items():
        ws = wb.create_sheet(title=sheet_name)
        columns = list(SHEET_COLUMNS.get(sheet_name, rows[0].keys() if rows else []))
        ws.append(columns)
        for row in rows:
            ws.append([row.get(col) for col in columns])

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ---------------------------------------------------------------------------
# Row factories (one per sheet, sensible defaults)
# ---------------------------------------------------------------------------


def make_setting_row(**overrides: object) -> dict:
    """Single Settings row."""
    return {"key": "usd_rub_rate", "value": "90.0", "description": "", **overrides}


def make_loan_row(**overrides: object) -> dict:
    """Single Loans row with minimal valid data."""
    return {
        "code": "test_loan",
        "creditor": "Test Bank",
        "name": "Test Loan",
        "loan_type": "credit",
        "payment_method": "annuity",
        "original_amount": 100000,
        "interest_rate": 12.0,
        "opening_date": "2025-01-15",
        "closing_date": "2028-01-15",
        "prepayment_strategy": "reduce_payment",
        "priority": 5,
        "status": "active",
        "contract_number": "CN-001",
        "notes": "",
        **overrides,
    }


def make_balance_row(**overrides: object) -> dict:
    """Single Balances row."""
    return {
        "loan_code": "test_loan",
        "snapshot_date": "2025-06-01",
        "current_balance": 95000,
        "principal_balance": 93000,
        "accrued_interest": 2000,
        "source": "imported",
        "notes": "",
        **overrides,
    }


def make_schedule_row(**overrides: object) -> dict:
    """Single Schedule row."""
    return {
        "loan_code": "test_loan",
        "due_date": "2025-07-15",
        "amount": 3500,
        "principal_part": 2800,
        "interest_part": 700,
        "accuracy": "estimate",
        "can_pay_early": True,
        "income_code": "",
        "notes": "",
        **overrides,
    }


def make_income_row(**overrides: object) -> dict:
    """Single Incomes row."""
    return {
        "code": "salary_2025_07_10",
        "expected_date": "2025-07-10",
        "amount_rub": 45000,
        "amount_usd": None,
        "name": "Зарплата",
        "status": "expected",
        "notes": "",
        **overrides,
    }


def make_actual_payment_row(**overrides: object) -> dict:
    """Single ActualPayments row."""
    return {
        "loan_code": "test_loan",
        "payment_date": "2025-06-15",
        "amount": 3500,
        "principal_part": 2800,
        "interest_part": 700,
        "payment_type": "regular",
        "planned_due_date": "2025-06-15",
        "notes": "",
        **overrides,
    }


# ---------------------------------------------------------------------------
# Ready-made workbooks
# ---------------------------------------------------------------------------


def minimal_valid_workbook() -> bytes:
    """Minimal happy-path workbook: Settings + Loans + Balances (1 row each)."""
    return build_workbook({
        "Settings": [make_setting_row()],
        "Loans": [make_loan_row()],
        "Balances": [make_balance_row()],
    })
