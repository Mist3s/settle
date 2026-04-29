"""ExportService — export current DB state to XLSX.

Produces a file structurally compatible with the import template
(architecture.md §11.7). Supports incremental export via ``since`` filter.
"""

from __future__ import annotations

import io
import uuid
from datetime import date
from decimal import Decimal
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Font
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.domain.models.balance import LoanBalance
from app.domain.models.income import Income
from app.domain.models.loan import Loan
from app.domain.models.payment import ActualPayment, PlannedPayment
from app.domain.models.settings import Setting

_BOLD = Font(bold=True)


def _cell_value(v: Any) -> Any:
    """Convert Python value to an Excel-friendly representation."""
    if v is None:
        return ""
    if isinstance(v, Decimal):
        return str(v)
    if isinstance(v, date):
        return v.isoformat()
    if isinstance(v, uuid.UUID):
        return str(v)
    if hasattr(v, "value"):
        return v.value
    return v


def _write_sheet(
    wb: Workbook,
    name: str,
    columns: list[str],
    rows: list[list[Any]],
) -> None:
    ws = wb.create_sheet(title=name)
    for col_idx, col_name in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = _BOLD
    for row_idx, row_data in enumerate(rows, start=2):
        for col_idx, value in enumerate(row_data, start=1):
            ws.cell(row=row_idx, column=col_idx, value=_cell_value(value))


async def _load_loan_code_map(
    session: AsyncSession,
    user_id: uuid.UUID,
) -> dict[uuid.UUID, str]:
    """Build loan_id → code mapping for FK resolution."""
    stmt = (
        select(Loan.id, Loan.code)
        .where(Loan.user_id == user_id, Loan.is_deleted.is_(False))
    )
    result = await session.execute(stmt)
    return {row.id: row.code for row in result}


async def _load_income_code_map(
    session: AsyncSession,
    user_id: uuid.UUID,
) -> dict[uuid.UUID, str]:
    stmt = (
        select(Income.id, Income.code)
        .where(Income.user_id == user_id, Income.is_deleted.is_(False))
    )
    result = await session.execute(stmt)
    return {row.id: row.code for row in result}


async def export_to_xlsx(
    session: AsyncSession,
    user_id: uuid.UUID,
    *,
    since: date | None = None,
) -> bytes:
    """Export all user data to XLSX.

    Parameters
    ----------
    since : date | None
        If provided, only records updated after this date are included
        (based on ``updated_at`` / ``created_at``).

    Returns
    -------
    bytes
        Raw XLSX content.
    """
    wb = Workbook()
    wb.remove(wb.active)  # type: ignore[arg-type]

    loan_code_map = await _load_loan_code_map(session, user_id)
    income_code_map = await _load_income_code_map(session, user_id)

    # --- Settings ---
    stmt = select(Setting).where(Setting.user_id == user_id)
    if since:
        stmt = stmt.where(Setting.updated_at >= since)
    settings = (await session.execute(stmt)).scalars().all()
    _write_sheet(wb, "Settings",
                 ["key", "value", "description"],
                 [[s.key, s.value, s.description] for s in settings])

    # --- Loans ---
    stmt = select(Loan).where(Loan.user_id == user_id, Loan.is_deleted.is_(False))
    if since:
        stmt = stmt.where(Loan.updated_at >= since)
    loans = (await session.execute(stmt)).scalars().all()
    _write_sheet(wb, "Loans",
                 ["code", "creditor", "name", "loan_type", "payment_method",
                  "original_amount", "interest_rate", "opening_date", "closing_date",
                  "prepayment_strategy", "priority", "status", "contract_number", "notes"],
                 [[l.code, l.creditor, l.name, l.loan_type, l.payment_method,
                   l.original_amount, l.interest_rate, l.opening_date, l.closing_date,
                   l.prepayment_strategy, l.priority, l.status, l.contract_number, l.notes]
                  for l in loans])

    # --- Balances ---
    loan_ids = list(loan_code_map.keys())
    if loan_ids:
        stmt = select(LoanBalance).where(LoanBalance.loan_id.in_(loan_ids))
        if since:
            stmt = stmt.where(LoanBalance.updated_at >= since)
        balances = (await session.execute(stmt)).scalars().all()
    else:
        balances = []
    _write_sheet(wb, "Balances",
                 ["loan_code", "snapshot_date", "current_balance",
                  "principal_balance", "accrued_interest", "source", "notes"],
                 [[loan_code_map.get(b.loan_id, ""), b.snapshot_date,
                   b.current_balance, b.principal_balance, b.accrued_interest,
                   b.source, b.notes] for b in balances])

    # --- Schedule (PlannedPayments) ---
    if loan_ids:
        stmt = (
            select(PlannedPayment)
            .where(PlannedPayment.user_id == user_id,
                   PlannedPayment.is_deleted.is_(False))
        )
        if since:
            stmt = stmt.where(PlannedPayment.updated_at >= since)
        planned = (await session.execute(stmt)).scalars().all()
    else:
        planned = []
    _write_sheet(wb, "Schedule",
                 ["loan_code", "due_date", "amount", "principal_part",
                  "interest_part", "accuracy", "can_pay_early", "income_code", "notes"],
                 [[loan_code_map.get(p.loan_id, ""), p.due_date,
                   p.amount, p.principal_part, p.interest_part,
                   p.accuracy, p.can_pay_early,
                   income_code_map.get(p.income_id, "") if p.income_id else "",
                   p.notes] for p in planned])

    # --- Incomes ---
    stmt = select(Income).where(Income.user_id == user_id, Income.is_deleted.is_(False))
    if since:
        stmt = stmt.where(Income.updated_at >= since)
    incomes = (await session.execute(stmt)).scalars().all()
    _write_sheet(wb, "Incomes",
                 ["code", "expected_date", "amount_rub", "amount_usd",
                  "name", "status", "notes"],
                 [[i.code, i.expected_date, i.amount, None,
                   i.name, i.status, i.notes] for i in incomes])

    # --- ActualPayments ---
    if loan_ids:
        stmt = select(ActualPayment).where(ActualPayment.loan_id.in_(loan_ids))
        if since:
            stmt = stmt.where(ActualPayment.updated_at >= since)
        actuals = (await session.execute(stmt)).scalars().all()
    else:
        actuals = []

    # Resolve planned_due_date from planned_payment_id
    planned_due_dates: dict[uuid.UUID, date] = {}
    if actuals:
        pp_ids = [a.planned_payment_id for a in actuals if a.planned_payment_id]
        if pp_ids:
            stmt2 = select(PlannedPayment.id, PlannedPayment.due_date).where(
                PlannedPayment.id.in_(pp_ids)
            )
            for row in (await session.execute(stmt2)):
                planned_due_dates[row.id] = row.due_date

    _write_sheet(wb, "ActualPayments",
                 ["loan_code", "payment_date", "amount", "principal_part",
                  "interest_part", "payment_type", "planned_due_date", "notes"],
                 [[loan_code_map.get(a.loan_id, ""), a.payment_date,
                   a.amount, a.principal_part, a.interest_part,
                   a.payment_type,
                   planned_due_dates.get(a.planned_payment_id) if a.planned_payment_id else None,
                   a.notes] for a in actuals])

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
