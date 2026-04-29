"""TemplateService — XLSX template generation (empty + with examples).

Generates a template file matching the import specification from
architecture.md §11.2. Two modes:
  - empty: headers only
  - with_examples: one example row per sheet, highlighted yellow,
    first cell prefixed with '[delete this row before import]'
"""

from __future__ import annotations

import io

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

from app.domain.constants.import_export import (
    ACTUAL_PAYMENTS_COLUMNS,
    BALANCES_COLUMNS,
    EXAMPLE_ROW_MARKER,
    INCOMES_COLUMNS,
    LOANS_COLUMNS,
    SCHEDULE_COLUMNS,
    SETTINGS_COLUMNS,
    SHEET_ACTUAL_PAYMENTS,
    SHEET_BALANCES,
    SHEET_INCOMES,
    SHEET_LOANS,
    SHEET_SCHEDULE,
    SHEET_SETTINGS,
)

# ---------------------------------------------------------------------------
# Styling
# ---------------------------------------------------------------------------

_YELLOW_FILL = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
_BOLD_FONT = Font(bold=True)

# ---------------------------------------------------------------------------
# Sheet definitions: (sheet_name, columns, example_row)
# Columns come from shared constants; example rows are template-specific.
# ---------------------------------------------------------------------------

_EXAMPLE_ROWS: dict[str, list[str]] = {
    SHEET_SETTINGS: [
        f"{EXAMPLE_ROW_MARKER} usd_rub_rate", "92.50", "Курс USD/RUB",  # noqa: RUF001
    ],
    SHEET_LOANS: [
        f"{EXAMPLE_ROW_MARKER} example_loan", "Сбербанк",
        "Потребительский кредит", "credit", "annuity", "500000.00",
        "12.50", "2025-01-15", "2028-01-15", "reduce_payment", "1",
        "active", "1234-5678", "Пример строки",
    ],
    SHEET_BALANCES: [
        f"{EXAMPLE_ROW_MARKER} example_loan", "2026-04-01",
        "450000.00", "445000.00", "5000.00", "imported", "",
    ],
    SHEET_SCHEDULE: [
        f"{EXAMPLE_ROW_MARKER} example_loan", "2026-05-15",
        "18602.00", "13900.00", "4702.00", "exact_contract", "true",
        "salary_2026_05_10", "",
    ],
    SHEET_INCOMES: [
        f"{EXAMPLE_ROW_MARKER} salary_2026_05_10", "2026-05-10",
        "150000.00", "", "Зарплата 10 мая", "expected", "",
    ],
    SHEET_ACTUAL_PAYMENTS: [
        f"{EXAMPLE_ROW_MARKER} example_loan", "2026-04-15",
        "18602.00", "13900.00", "4702.00", "regular", "2026-04-15", "",
    ],
}

_SHEET_DEFS: list[tuple[str, list[str], list[str]]] = [
    (SHEET_SETTINGS, SETTINGS_COLUMNS, _EXAMPLE_ROWS[SHEET_SETTINGS]),
    (SHEET_LOANS, LOANS_COLUMNS, _EXAMPLE_ROWS[SHEET_LOANS]),
    (SHEET_BALANCES, BALANCES_COLUMNS, _EXAMPLE_ROWS[SHEET_BALANCES]),
    (SHEET_SCHEDULE, SCHEDULE_COLUMNS, _EXAMPLE_ROWS[SHEET_SCHEDULE]),
    (SHEET_INCOMES, INCOMES_COLUMNS, _EXAMPLE_ROWS[SHEET_INCOMES]),
    (SHEET_ACTUAL_PAYMENTS, ACTUAL_PAYMENTS_COLUMNS, _EXAMPLE_ROWS[SHEET_ACTUAL_PAYMENTS]),
]


def generate_template(*, with_examples: bool = False) -> bytes:
    """Generate an XLSX template.

    Returns raw bytes suitable for writing to file or HTTP response.
    """
    wb = Workbook()
    # Remove default sheet created by openpyxl
    wb.remove(wb.active)  # type: ignore[arg-type]

    for sheet_name, columns, example_row in _SHEET_DEFS:
        ws = wb.create_sheet(title=sheet_name)

        # Header row
        for col_idx, col_name in enumerate(columns, start=1):
            cell = ws.cell(row=1, column=col_idx, value=col_name)
            cell.font = _BOLD_FONT

        # Example row (optional)
        if with_examples:
            for col_idx, value in enumerate(example_row, start=1):
                cell = ws.cell(row=2, column=col_idx, value=value)
                cell.fill = _YELLOW_FILL

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
