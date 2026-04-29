"""Parse XLSX workbook bytes into typed DTOs.

Reads sheets recognised by ``SHEET_COLUMNS``, validates headers via
``header_validator``, parses each data row through the corresponding
Pydantic DTO, skips blank rows and example-marker rows
(``[delete this row before import]``), and collects all errors with
sheet / row / column coordinates.
"""

from __future__ import annotations

from io import BytesIO
from typing import Any

from openpyxl import load_workbook
from pydantic import BaseModel, ValidationError

from app.domain.constants.import_export import (
    EXAMPLE_ROW_MARKER,
    SHEET_COLUMNS,
)
from app.domain.schemas.import_dto import (
    ActualPaymentImportRow,
    BalanceImportRow,
    IncomeImportRow,
    LoanImportRow,
    ScheduleImportRow,
    SettingImportRow,
)
from app.domain.schemas.import_report import (
    ImportError as ImportErr,
)
from app.domain.schemas.import_report import (
    ImportWarning,
)
from app.services.import_.header_validator import (
    validate_headers,
    validate_required_sheets,
)

# ---------------------------------------------------------------------------
# Sheet → DTO mapping
# ---------------------------------------------------------------------------

_SHEET_DTO: dict[str, type[BaseModel]] = {
    "Settings": SettingImportRow,
    "Loans": LoanImportRow,
    "Balances": BalanceImportRow,
    "Schedule": ScheduleImportRow,
    "Incomes": IncomeImportRow,
    "ActualPayments": ActualPaymentImportRow,
}




# ---------------------------------------------------------------------------
# Result container
# ---------------------------------------------------------------------------

class ParsedData(BaseModel):
    """Result of workbook parsing: dict[sheet_name, list[DTO]]."""

    settings: list[SettingImportRow] = []
    loans: list[LoanImportRow] = []
    balances: list[BalanceImportRow] = []
    schedule: list[ScheduleImportRow] = []
    incomes: list[IncomeImportRow] = []
    actual_payments: list[ActualPaymentImportRow] = []


_ATTR_BY_SHEET: dict[str, str] = {
    "Settings": "settings",
    "Loans": "loans",
    "Balances": "balances",
    "Schedule": "schedule",
    "Incomes": "incomes",
    "ActualPayments": "actual_payments",
}


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _is_blank_row(values: list[Any]) -> bool:
    """Return *True* if every cell in the row is empty / None."""
    return all(v is None or (isinstance(v, str) and v.strip() == "") for v in values)


def _is_example_row(values: list[Any]) -> bool:
    """Return *True* if any cell contains the example marker."""
    return any(
        isinstance(v, str) and EXAMPLE_ROW_MARKER in v
        for v in values
    )


def _pydantic_errors_to_import_errors(
    exc: ValidationError,
    sheet: str,
    row_idx: int,
    columns: list[str],
) -> list[ImportErr]:
    """Convert Pydantic ``ValidationError`` into ``ImportError`` list."""
    result: list[ImportErr] = []
    for err in exc.errors():
        # loc is a tuple — first element is the field name
        field = str(err["loc"][0]) if err["loc"] else "?"
        col_name = field
        result.append(
            ImportErr(
                sheet=sheet,
                row=row_idx,
                column=col_name,
                message=err["msg"],
            ),
        )
    return result


# ---------------------------------------------------------------------------
# Public API
# ---------------------------------------------------------------------------

def parse_workbook(
    file_bytes: bytes,
) -> tuple[ParsedData, list[ImportErr], list[ImportWarning]]:
    """Parse *file_bytes* (XLSX) into :class:`ParsedData`.

    Returns ``(parsed_data, errors, warnings)``.

    * Headers are validated per-sheet via :func:`validate_headers`.
    * Required-sheet presence is validated via :func:`validate_required_sheets`.
    * Sheets not in ``SHEET_COLUMNS`` produce a warning, not an error.
    * Blank rows and example-marker rows are silently skipped.
    * Pydantic validation errors are collected with ``(sheet, row, column)``
      coordinates — parsing continues on error.
    """
    wb = load_workbook(BytesIO(file_bytes), read_only=True, data_only=True)

    errors: list[ImportErr] = []
    warnings: list[ImportWarning] = []
    parsed = ParsedData()

    # --- required sheets ---
    errors.extend(validate_required_sheets(wb.sheetnames))

    for ws in wb.worksheets:
        sheet_name: str = ws.title

        # Unknown sheet → warning, skip
        if sheet_name not in SHEET_COLUMNS:
            warnings.append(
                ImportWarning(
                    sheet=sheet_name,
                    message=f"Неизвестный лист «{sheet_name}» — пропущен",
                ),
            )
            continue

        rows_iter = ws.iter_rows(values_only=True)

        # --- header row ---
        header_raw = next(rows_iter, None)
        if header_raw is None:
            errors.append(
                ImportErr(sheet=sheet_name, message="Лист пуст (нет строки заголовков)"),
            )
            continue

        columns = [str(c).strip() if c is not None else "" for c in header_raw]
        header_errors = validate_headers(sheet_name, columns)
        if header_errors:
            errors.extend(header_errors)
            continue  # cannot parse rows when headers are wrong

        dto_cls = _SHEET_DTO[sheet_name]
        attr = _ATTR_BY_SHEET[sheet_name]
        target_list: list[Any] = getattr(parsed, attr)

        # --- data rows (1-indexed: header is row 1) ---
        for row_num, row_values in enumerate(rows_iter, start=2):
            values = list(row_values)

            if _is_blank_row(values):
                continue
            if _is_example_row(values):
                continue

            # Build dict: column_name → cell value
            row_dict: dict[str, Any] = {}
            for col_name, cell_val in zip(columns, values, strict=False):
                row_dict[col_name] = cell_val

            try:
                dto = dto_cls.model_validate(row_dict)
            except ValidationError as exc:
                errors.extend(
                    _pydantic_errors_to_import_errors(exc, sheet_name, row_num, columns),
                )
                continue

            target_list.append(dto)

    wb.close()
    return parsed, errors, warnings
