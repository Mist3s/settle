"""Integration tests for import/export REST endpoints (architecture.md §8.2).

Tests exercise the full HTTP layer via AsyncClient with overridden
``get_session`` and ``get_current_user`` dependencies.
"""

from __future__ import annotations

from io import BytesIO
from uuid import uuid4

import pytest
from httpx import ASGITransport, AsyncClient
from openpyxl import load_workbook
from sqlalchemy.ext.asyncio import AsyncSession

from app.api.deps import get_current_user
from app.core.database import get_session
from app.core.security import hash_password
from app.domain.models.user import User
from app.main import app
from tests.fixtures.import_fixtures import minimal_valid_workbook

# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

_XLSX_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


async def _create_user(session: AsyncSession) -> User:
    user = User(email="api_test@settle.local", password_hash=hash_password("pw"))
    session.add(user)
    await session.flush()
    await session.refresh(user)
    return user


@pytest.fixture()
async def auth_client(db_session: AsyncSession):
    """AsyncClient with overridden DB session *and* current_user."""
    user = await _create_user(db_session)

    async def _override_session():
        yield db_session

    def _override_user():
        return user

    app.dependency_overrides[get_session] = _override_session
    app.dependency_overrides[get_current_user] = _override_user

    async with AsyncClient(
        transport=ASGITransport(app=app),
        base_url="http://test",
    ) as ac:
        yield ac

    app.dependency_overrides.clear()


# ---------------------------------------------------------------------------
# POST /api/import/excel — multipart upload → DryRunReport
# ---------------------------------------------------------------------------


@pytest.mark.asyncio()
async def test_upload_xlsx_returns_dry_run_report(auth_client: AsyncClient):
    """Given: valid XLSX, When: POST /api/import/excel, Then: 200 + report."""
    xlsx_bytes = minimal_valid_workbook()
    resp = await auth_client.post(
        "/api/import/excel",
        files={"file": ("test.xlsx", xlsx_bytes, _XLSX_MIME)},
    )
    assert resp.status_code == 200
    body = resp.json()
    assert "import_id" in body
    assert "expires_at" in body


# ---------------------------------------------------------------------------
# POST /api/import/excel/commit — valid + expired/missing import_id
# ---------------------------------------------------------------------------


@pytest.mark.asyncio()
async def test_commit_valid_import(auth_client: AsyncClient):
    """Given: dry-run done, When: commit with import_id, Then: 200."""
    xlsx_bytes = minimal_valid_workbook()
    dry = await auth_client.post(
        "/api/import/excel",
        files={"file": ("test.xlsx", xlsx_bytes, _XLSX_MIME)},
    )
    import_id = dry.json()["import_id"]

    resp = await auth_client.post(
        "/api/import/excel/commit",
        json={"import_id": import_id},
    )
    assert resp.status_code == 200
    body = resp.json()
    assert "loans_created" in body


@pytest.mark.asyncio()
async def test_commit_missing_import_id_returns_410(auth_client: AsyncClient):
    """Given: non-existent import_id, When: commit, Then: 410 Gone."""
    resp = await auth_client.post(
        "/api/import/excel/commit",
        json={"import_id": str(uuid4())},
    )
    assert resp.status_code == 410


# ---------------------------------------------------------------------------
# GET /api/import/template — empty + with_examples
# ---------------------------------------------------------------------------


@pytest.mark.asyncio()
async def test_template_download_empty(auth_client: AsyncClient):
    """Given: no params, When: GET /api/import/template, Then: valid XLSX."""
    resp = await auth_client.get("/api/import/template")
    assert resp.status_code == 200
    assert _XLSX_MIME in resp.headers["content-type"]

    wb = load_workbook(BytesIO(resp.content))
    assert "Loans" in wb.sheetnames
    wb.close()


@pytest.mark.asyncio()
async def test_template_download_with_examples(auth_client: AsyncClient):
    """Given: with_examples=true, When: GET template, Then: XLSX with data rows."""
    resp = await auth_client.get("/api/import/template?with_examples=true")
    assert resp.status_code == 200

    wb = load_workbook(BytesIO(resp.content))
    loans_ws = wb["Loans"]
    # Header + at least one example row.
    assert loans_ws.max_row >= 2
    wb.close()


# ---------------------------------------------------------------------------
# GET /api/export/excel — download + since filter
# ---------------------------------------------------------------------------


@pytest.mark.asyncio()
async def test_export_xlsx(auth_client: AsyncClient):
    """Given: data in DB, When: GET /api/export/excel, Then: valid XLSX."""
    # Seed some data via import first.
    xlsx_bytes = minimal_valid_workbook()
    dry = await auth_client.post(
        "/api/import/excel",
        files={"file": ("test.xlsx", xlsx_bytes, _XLSX_MIME)},
    )
    await auth_client.post(
        "/api/import/excel/commit",
        json={"import_id": dry.json()["import_id"]},
    )

    resp = await auth_client.get("/api/export/excel")
    assert resp.status_code == 200
    assert _XLSX_MIME in resp.headers["content-type"]

    wb = load_workbook(BytesIO(resp.content))
    assert "Loans" in wb.sheetnames
    wb.close()


@pytest.mark.asyncio()
async def test_export_since_filter(auth_client: AsyncClient):
    """Given: since far in future, When: GET /api/export/excel?since=, Then: XLSX (empty data)."""
    resp = await auth_client.get("/api/export/excel?since=2099-01-01")
    assert resp.status_code == 200
    assert _XLSX_MIME in resp.headers["content-type"]


# ---------------------------------------------------------------------------
# Unauthenticated → 401/403
# ---------------------------------------------------------------------------


@pytest.mark.asyncio()
async def test_unauthenticated_request_rejected(db_session: AsyncSession):
    """Given: no auth, When: GET /api/import/template, Then: 401 or 403."""
    async def _override_session():
        yield db_session

    app.dependency_overrides[get_session] = _override_session

    try:
        async with AsyncClient(
            transport=ASGITransport(app=app),
            base_url="http://test",
        ) as ac:
            resp = await ac.get("/api/import/template")
            assert resp.status_code in (401, 403)
    finally:
        app.dependency_overrides.clear()


# ---------------------------------------------------------------------------
# Non-XLSX file → 400
# ---------------------------------------------------------------------------


@pytest.mark.asyncio()
async def test_non_xlsx_file_returns_400(auth_client: AsyncClient):
    """Given: JPEG bytes, When: POST /api/import/excel, Then: 400."""
    fake_jpeg = b"\xff\xd8\xff\xe0" + b"\x00" * 100  # JPEG header
    resp = await auth_client.post(
        "/api/import/excel",
        files={"file": ("photo.jpg", fake_jpeg, "image/jpeg")},
    )
    assert resp.status_code == 400
    assert "XLSX" in resp.json()["detail"]
