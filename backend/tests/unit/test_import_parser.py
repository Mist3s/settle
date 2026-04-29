"""Unit tests for services/import_/parser.py."""

from __future__ import annotations

from io import BytesIO

from openpyxl import Workbook

from app.domain.constants.import_export import EXAMPLE_ROW_MARKER, SHEET_COLUMNS
from app.services.import_.parser import (
    ParsedData,
    parse_workbook,
)
from tests.fixtures.import_fixtures import (
    build_workbook,
    make_actual_payment_row,
    make_balance_row,
    make_income_row,
    make_loan_row,
    make_schedule_row,
    make_setting_row,
    minimal_valid_workbook,
)


class TestParseWorkbookHappyPath:
    """Minimal valid workbook is parsed without errors."""

    def test_happy_path_three_sheets(self) -> None:
        wb_bytes = minimal_valid_workbook()
        parsed, errors, warnings = parse_workbook(wb_bytes)

        assert errors == [], f"Unexpected errors: {errors}"
        assert warnings == []
        assert isinstance(parsed, ParsedData)
        assert len(parsed.settings) == 1
        assert len(parsed.loans) == 1
        assert len(parsed.balances) == 1
        assert parsed.loans[0].code == "test_loan"

    def test_all_six_sheets(self) -> None:
        wb_bytes = build_workbook({
            "Settings": [make_setting_row()],
            "Loans": [make_loan_row()],
            "Balances": [make_balance_row()],
            "Schedule": [make_schedule_row()],
            "Incomes": [make_income_row()],
            "ActualPayments": [make_actual_payment_row()],
        })

        parsed, errors, warnings = parse_workbook(wb_bytes)
        assert errors == [], f"Unexpected errors: {errors}"
        assert warnings == []
        assert len(parsed.schedule) == 1
        assert len(parsed.incomes) == 1
        assert len(parsed.actual_payments) == 1


class TestExampleRowSkip:
    """Rows with the example marker in the first cell are skipped."""

    def test_example_row_skipped(self) -> None:
        marker = EXAMPLE_ROW_MARKER
        example_loan = make_loan_row(code=f"{marker}")
        real_loan = make_loan_row(code="real_loan")

        wb_bytes = build_workbook({
            "Settings": [make_setting_row()],
            "Loans": [example_loan, real_loan],
            "Balances": [make_balance_row(loan_code="real_loan")],
        })

        parsed, errors, _warnings = parse_workbook(wb_bytes)
        assert errors == [], f"Unexpected errors: {errors}"
        assert len(parsed.loans) == 1
        assert parsed.loans[0].code == "real_loan"


class TestBlankRowSkip:
    """Completely blank rows are silently skipped."""

    def test_blank_rows_ignored(self) -> None:
        wb = Workbook()
        wb.remove(wb.active)  # type: ignore[arg-type]

        # Settings — with a blank row between header and data
        ws = wb.create_sheet("Settings")
        cols = list(SHEET_COLUMNS["Settings"])
        ws.append(cols)
        ws.append([None] * len(cols))  # blank
        setting = make_setting_row()
        ws.append([setting.get(c) for c in cols])

        # Loans
        ws2 = wb.create_sheet("Loans")
        cols2 = list(SHEET_COLUMNS["Loans"])
        ws2.append(cols2)
        row_data = make_loan_row()
        ws2.append([row_data.get(c) for c in cols2])

        # Balances
        ws3 = wb.create_sheet("Balances")
        cols3 = list(SHEET_COLUMNS["Balances"])
        ws3.append(cols3)
        bal = make_balance_row()
        ws3.append([bal.get(c) for c in cols3])

        buf = BytesIO()
        wb.save(buf)
        wb_bytes = buf.getvalue()

        parsed, errors, _warnings = parse_workbook(wb_bytes)
        assert errors == [], f"Unexpected errors: {errors}"
        assert len(parsed.settings) == 1


class TestTypeErrorCoordinates:
    """Pydantic validation errors include sheet, row, and column."""

    def test_invalid_decimal_reports_coordinates(self) -> None:
        bad_loan = make_loan_row(original_amount="not_a_number")

        wb_bytes = build_workbook({
            "Settings": [make_setting_row()],
            "Loans": [bad_loan],
            "Balances": [make_balance_row()],
        })

        parsed, errors, _warnings = parse_workbook(wb_bytes)
        assert len(errors) >= 1
        err = errors[0]
        assert err.sheet == "Loans"
        assert err.row == 2  # first data row
        assert err.column == "original_amount"
        # The row with an error should not appear in parsed data
        assert len(parsed.loans) == 0

    def test_invalid_enum_reports_coordinates(self) -> None:
        bad_loan = make_loan_row(loan_type="totally_invalid_type")

        wb_bytes = build_workbook({
            "Settings": [make_setting_row()],
            "Loans": [bad_loan],
            "Balances": [make_balance_row()],
        })

        _parsed, errors, _warnings = parse_workbook(wb_bytes)
        assert len(errors) >= 1
        err = errors[0]
        assert err.sheet == "Loans"
        assert err.row == 2
        assert err.column == "loan_type"


class TestEmptyWorkbook:
    """Workbook with required sheets but no data rows."""

    def test_empty_required_sheets(self) -> None:
        wb_bytes = build_workbook({
            "Settings": [],
            "Loans": [],
            "Balances": [],
        })

        parsed, errors, _warnings = parse_workbook(wb_bytes)
        assert errors == []
        assert len(parsed.settings) == 0
        assert len(parsed.loans) == 0
        assert len(parsed.balances) == 0


class TestMissingRequiredSheet:
    """Missing required sheet produces an error."""

    def test_missing_loans(self) -> None:
        wb_bytes = build_workbook({
            "Settings": [make_setting_row()],
            "Balances": [make_balance_row()],
        })

        _parsed, errors, _warnings = parse_workbook(wb_bytes)
        assert any("Loans" in e.sheet for e in errors)


class TestUnknownSheetWarning:
    """Unknown sheets produce a warning, not an error."""

    def test_unknown_sheet_produces_warning(self) -> None:
        wb = Workbook()
        wb.remove(wb.active)  # type: ignore[arg-type]

        # Required sheets
        for name in ("Settings", "Loans", "Balances"):
            ws = wb.create_sheet(name)
            cols = list(SHEET_COLUMNS[name])
            ws.append(cols)

        # Unknown sheet
        ws_unknown = wb.create_sheet("RandomStuff")
        ws_unknown.append(["a", "b", "c"])
        ws_unknown.append([1, 2, 3])

        buf = BytesIO()
        wb.save(buf)
        wb_bytes = buf.getvalue()

        _parsed, errors, warnings = parse_workbook(wb_bytes)
        # No errors from the unknown sheet
        assert errors == [], f"Unexpected errors: {errors}"
        # But a warning is produced
        assert len(warnings) == 1
        assert warnings[0].sheet == "RandomStuff"
        assert "пропущен" in warnings[0].message


class TestMultipleRows:
    """Multiple data rows are all parsed."""

    def test_two_loans_parsed(self) -> None:
        loan1 = make_loan_row(code="loan_a")
        loan2 = make_loan_row(code="loan_b")

        wb_bytes = build_workbook({
            "Settings": [make_setting_row()],
            "Loans": [loan1, loan2],
            "Balances": [
                make_balance_row(loan_code="loan_a"),
                make_balance_row(loan_code="loan_b"),
            ],
        })

        parsed, errors, _warnings = parse_workbook(wb_bytes)
        assert errors == []
        assert len(parsed.loans) == 2
        codes = {loan.code for loan in parsed.loans}
        assert codes == {"loan_a", "loan_b"}
